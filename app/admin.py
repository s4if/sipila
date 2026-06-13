from datetime import datetime, timezone

from flask import Blueprint, jsonify, request, session, url_for
from werkzeug.security import check_password_hash, generate_password_hash

from .db import db
from .forms import GuruForm, KategoriForm, RombelForm, SiswaForm
from .helper import (
    WIB,
    admin_required,
    hx_render,
    sanitize,
    superadmin_required,
)
from .models import (
    BorrowingRequest,
    Category,
    CategoryTeacher,
    ClassGroup,
    Student,
    Teacher,
)

bp = Blueprint("admin", __name__, url_prefix="/admin")


@bp.route("/")
@admin_required
def beranda():
    return hx_render("admin/beranda.jinja")


@bp.route("/ganti_password", methods=["GET", "POST"])
@admin_required
def ganti_password():
    if request.method == "GET":
        return hx_render("admin/ganti_password.jinja")

    notif = {}
    admin = Teacher.query.filter_by(username=session["admin_name"]).first()
    if request.form["new_password"] != request.form["confirm_password"]:
        notif["error"] = "Konfirmasi password tidak sesuai"
    elif admin and check_password_hash(
        admin.password, request.form["current_password"]
    ):
        admin.password = generate_password_hash(request.form["new_password"])
        db.session.commit()
        notif["success"] = "Password berhasil diubah"
    else:
        notif["error"] = "Password lama tidak sesuai"
    return hx_render("admin/ganti_password.jinja", **notif)


# ---- Rombel (ClassGroup) CRUD ----


@bp.route("/rombel")
@admin_required
def rombel():
    return hx_render("admin/rombel.jinja")


@bp.route("/rombel/data")
@admin_required
def rombel_data():
    from sqlalchemy import func
    from sqlalchemy.orm import joinedload

    class_groups = (
        ClassGroup.query.options(joinedload(ClassGroup.homeroom_teacher))
        .order_by(ClassGroup.id)
        .all()
    )

    active_counts = dict(
        db.session.query(Student.class_group_id, func.count(Student.id))
        .filter(~Student.is_deleted, Student.class_group_id.isnot(None))
        .group_by(Student.class_group_id)
        .all()
    )

    is_superadmin = session.get("is_superadmin", False)
    data = []
    for i, cg in enumerate(class_groups, 1):
        guru = (
            (cg.homeroom_teacher.name or "[nama belum di set]")
            if cg.homeroom_teacher
            else "-"
        )
        data.append(
            {
                "no": i,
                "display_name": cg.display_name,
                "grade_level": cg.grade_level,
                "major": cg.major or "-",
                "homeroom_teacher": guru,
                "student_count": active_counts.get(cg.id, 0),
                "actions": (
                    '<a class="btn btn-sm btn-warning" '
                    f'onclick="edit_rombel({cg.id})">'
                    '<i class="bi bi-pencil"></i> Edit</a> '
                    '<button type="button" class="btn btn-sm btn-danger" '
                    f"onclick=\"hapus_rombel({cg.id}, '{sanitize(cg.display_name)}')\">"
                    '<i class="bi bi-trash"></i> Hapus</button>'
                )
                if is_superadmin
                else "-",
            }
        )
    return jsonify(data=data)


@bp.route("/rombel/tambah", methods=["GET", "POST"])
@superadmin_required
def rombel_tambah():
    form = RombelForm()
    form.homeroom_teacher_id.choices = [("", "Belum ditentukan")] + [
        (a.id, a.name or "[nama belum di set]")
        for a in Teacher.query.order_by(Teacher.username).all()
    ]
    if request.method == "GET":
        return hx_render("admin/rombel_form.jinja", class_group=None, form=form)

    if not form.validate_on_submit():
        return hx_render("admin/rombel_form.jinja", class_group=None, form=form)

    notif = {}
    class_group = ClassGroup(
        name=sanitize(form.name.data),
        grade_level=form.grade_level.data,
        major=form.major.data or None,
        homeroom_teacher_id=form.homeroom_teacher_id.data or None,
    )
    db.session.add(class_group)
    db.session.commit()
    notif["success"] = "Rombel berhasil ditambahkan"
    return hx_render("admin/rombel.jinja", push_url="admin.rombel", **notif)


@bp.route("/rombel/edit/<int:id>", methods=["GET", "POST"])
@superadmin_required
def rombel_edit(id):
    class_group = ClassGroup.query.get_or_404(id)
    form = RombelForm(obj=class_group)
    form.homeroom_teacher_id.choices = [("", "Belum ditentukan")] + [
        (a.id, a.name or "[nama belum di set]")
        for a in Teacher.query.order_by(Teacher.username).all()
    ]
    if request.method == "GET":
        return hx_render("admin/rombel_form.jinja", class_group=class_group, form=form)

    if not form.validate_on_submit():
        return hx_render("admin/rombel_form.jinja", class_group=class_group, form=form)

    notif = {}
    class_group.name = sanitize(form.name.data)
    class_group.grade_level = form.grade_level.data
    class_group.major = form.major.data or None
    class_group.homeroom_teacher_id = form.homeroom_teacher_id.data or None
    db.session.commit()
    notif["success"] = "Rombel berhasil diperbarui"
    return hx_render("admin/rombel.jinja", push_url="admin.rombel", **notif)


@bp.route("/rombel/hapus", methods=["POST"])
@superadmin_required
def rombel_hapus():
    id = request.form.get("id", type=int)
    class_group = ClassGroup.query.get_or_404(id)
    notif = {}
    active_students = Student.query.filter_by(
        class_group_id=id, is_deleted=False
    ).first()
    if active_students:
        notif["error"] = "Rombel tidak dapat dihapus karena masih memiliki siswa aktif"
    else:
        db.session.delete(class_group)
        db.session.commit()
        notif["success"] = "Rombel berhasil dihapus"
    return hx_render("admin/rombel.jinja", push_url="admin.rombel", **notif)


# ---- Siswa (Student) CRUD ----


@bp.route("/siswa")
@admin_required
def siswa():
    return hx_render("admin/siswa.jinja")


@bp.route("/siswa/data")
@admin_required
def siswa_data():
    from sqlalchemy.orm import joinedload

    students = (
        Student.query.filter_by(is_deleted=False)
        .options(joinedload(Student.class_group))
        .order_by(Student.id)
        .all()
    )
    is_superadmin = session.get("is_superadmin", False)
    data = []
    for i, student in enumerate(students, 1):
        data.append(
            {
                "no": i,
                "student_id": student.student_id,
                "name": student.name,
                "class_group": student.class_group.display_name
                if student.class_group
                else "-",
                "admin_note": student.admin_note or "",
                "actions": (
                    '<a class="btn btn-sm btn-warning" '
                    f'onclick="edit_siswa({student.id})">'
                    '<i class="bi bi-pencil"></i> Edit</a> '
                    '<button type="button" class="btn btn-sm btn-danger" '
                    f"onclick=\"hapus_siswa({student.id}, '{sanitize(student.name)}')\">"
                    '<i class="bi bi-trash"></i> Hapus</button>'
                )
                if is_superadmin
                else "-",
            }
        )
    return jsonify(data=data)


@bp.route("/siswa/tambah", methods=["GET", "POST"])
@superadmin_required
def siswa_tambah():
    form = SiswaForm()
    form.class_group_id.choices = [("", "Pilih rombel")] + [
        (cg.id, cg.display_name)
        for cg in ClassGroup.query.order_by(ClassGroup.id).all()
    ]
    if request.method == "GET":
        return hx_render("admin/siswa_form.jinja", student=None, form=form)

    if not form.validate_on_submit():
        return hx_render("admin/siswa_form.jinja", student=None, form=form)

    notif = {}
    existing = Student.query.filter_by(student_id=form.student_id.data).first()
    if existing:
        notif["error"] = "NIS sudah terdaftar"
        return hx_render(
            "admin/siswa_form.jinja",
            student=None,
            form=form,
            **notif,
        )

    if not form.password.data:
        form.password.errors.append("Password wajib diisi")
        return hx_render("admin/siswa_form.jinja", student=None, form=form)

    student = Student(
        student_id=sanitize(form.student_id.data),
        name=sanitize(form.name.data),
        password=generate_password_hash(
            form.password.data, method="pbkdf2:sha256", salt_length=16
        ),
        class_group_id=form.class_group_id.data,
        admin_note=sanitize(form.admin_note.data),
    )
    db.session.add(student)
    db.session.commit()
    notif["success"] = "Siswa berhasil ditambahkan"
    return hx_render("admin/siswa.jinja", push_url="admin.siswa", **notif)


@bp.route("/siswa/edit/<int:id>", methods=["GET", "POST"])
@superadmin_required
def siswa_edit(id):
    student = Student.query.get_or_404(id)
    form = SiswaForm(obj=student)
    form.class_group_id.choices = [("", "Pilih rombel")] + [
        (cg.id, cg.display_name)
        for cg in ClassGroup.query.order_by(ClassGroup.id).all()
    ]
    if request.method == "GET":
        return hx_render("admin/siswa_form.jinja", student=student, form=form)

    if not form.validate_on_submit():
        return hx_render("admin/siswa_form.jinja", student=student, form=form)

    notif = {}
    existing = Student.query.filter(
        Student.student_id == form.student_id.data,
        Student.id != id,
    ).first()
    if existing:
        notif["error"] = "NIS sudah digunakan siswa lain"
        return hx_render(
            "admin/siswa_form.jinja",
            student=student,
            form=form,
            **notif,
        )

    student.student_id = sanitize(form.student_id.data)
    student.name = sanitize(form.name.data)
    if form.password.data:
        student.password = generate_password_hash(
            form.password.data, method="pbkdf2:sha256", salt_length=16
        )
    student.class_group_id = form.class_group_id.data
    student.admin_note = sanitize(form.admin_note.data)
    db.session.commit()
    notif["success"] = "Siswa berhasil diperbarui"
    return hx_render("admin/siswa.jinja", push_url="admin.siswa", **notif)


@bp.route("/siswa/hapus", methods=["POST"])
@superadmin_required
def siswa_hapus():
    id = request.form.get("id", type=int)
    student = Student.query.get_or_404(id)
    student.is_deleted = True
    db.session.commit()
    notif = {"success": "Siswa berhasil dihapus"}
    return hx_render("admin/siswa.jinja", push_url="admin.siswa", **notif)


@bp.route("/siswa/template")
@superadmin_required
def siswa_template():
    from io import BytesIO

    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Siswa"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    label_font = Font(bold=True)

    ws["A1"] = "Mode"
    ws["A1"].font = label_font
    ws["A2"] = "Jumlah"
    ws["A2"].font = label_font

    ws["H1"] = "Nama Rombel"
    ws["H1"].font = header_font
    ws["H1"].fill = header_fill
    ws["I1"] = "id"
    ws["I1"].font = header_font
    ws["I1"].fill = header_fill

    class_groups = ClassGroup.query.order_by(ClassGroup.id).all()
    for i, cg in enumerate(class_groups, start=2):
        ws.cell(row=i, column=8, value=cg.display_name)
        ws.cell(row=i, column=9, value=cg.id)

    header_row = 4
    headers = ["NIS", "Nama", "Password", "Rombel", "Catatan Admin"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["H"].width = 25
    ws.column_dimensions["I"].width = 8

    dv_mode = DataValidation(
        type="list",
        formula1='"skip,update"',
        allow_blank=False,
    )
    dv_mode.error = "Mode harus 'skip' atau 'update'"
    dv_mode.errorTitle = "Mode Tidak Valid"
    dv_mode.prompt = "Pilih 'skip' atau 'update'"
    dv_mode.promptTitle = "Mode Import"
    dv_mode.add("B1")
    ws.add_data_validation(dv_mode)

    dv_jumlah = DataValidation(
        type="whole",
        operator="greaterThan",
        formula1=0,
        allow_blank=False,
    )
    dv_jumlah.error = "Jumlah harus berupa angka lebih dari 0"
    dv_jumlah.errorTitle = "Jumlah Tidak Valid"
    dv_jumlah.prompt = "Masukkan jumlah siswa yang akan diimport"
    dv_jumlah.promptTitle = "Jumlah Siswa"
    dv_jumlah.add("B2")
    ws.add_data_validation(dv_jumlah)

    if class_groups:
        cg_ids = ",".join(str(cg.id) for cg in class_groups)
        dv_rombel = DataValidation(
            type="list",
            formula1='"{}"'.format(cg_ids),
            allow_blank=True,
        )
        dv_rombel.error = "Pilih id rombel dari daftar yang tersedia"
        dv_rombel.errorTitle = "Rombel Tidak Valid"
        dv_rombel.prompt = "Pilih id rombel dari daftar (lihat kolom H-I)"
        dv_rombel.promptTitle = "Rombel"
        dv_rombel.add("D5:D504")
        ws.add_data_validation(dv_rombel)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="template_import_siswa.xlsx",
    )


@bp.route("/siswa/import", methods=["POST"])
@superadmin_required
def siswa_import():
    import openpyxl

    notif = {}

    if "file" not in request.files:
        notif["error"] = "Tidak ada file yang diunggah"
        return hx_render("admin/siswa.jinja", push_url="admin.siswa", **notif)

    file = request.files["file"]
    if not file.filename or not file.filename.endswith(".xlsx"):
        notif["error"] = "File harus berformat .xlsx"
        return hx_render("admin/siswa.jinja", push_url="admin.siswa", **notif)

    try:
        wb = openpyxl.load_workbook(file)
    except Exception:
        notif["error"] = "File tidak dapat dibaca. Pastikan file xlsx valid."
        return hx_render("admin/siswa.jinja", push_url="admin.siswa", **notif)

    ws = wb.active

    mode = str(ws["B1"].value or "skip").strip().lower()
    if mode not in ("skip", "update"):
        notif["error"] = "Mode harus 'skip' atau 'update' (sel B1)"
        return hx_render("admin/siswa.jinja", push_url="admin.siswa", **notif)

    rombel_map = {}
    row = 2
    while ws.cell(row=row, column=8).value is not None:
        cg_id = ws.cell(row=row, column=9).value
        if cg_id is not None:
            try:
                rombel_map[int(cg_id)] = ws.cell(row=row, column=8).value
            except (ValueError, TypeError):
                notif["error"] = "Rombel id tidak valid pada baris {}: '{}'".format(row, cg_id)
                return hx_render("admin/siswa.jinja", push_url="admin.siswa", **notif)
        row += 1

    valid_cg_ids = {cg.id for cg in ClassGroup.query.all()}
    invalid_rombel_ids = set(rombel_map.keys()) - valid_cg_ids
    if invalid_rombel_ids:
        notif["error"] = "Rombel id tidak valid: {}".format(
            ", ".join(str(i) for i in sorted(invalid_rombel_ids))
        )
        return hx_render("admin/siswa.jinja", push_url="admin.siswa", **notif)

    jumlah_siswa = ws["B2"].value
    try:
        jumlah_siswa = int(jumlah_siswa)
    except (ValueError, TypeError):
        notif["error"] = "Jumlah siswa (sel B2) harus berupa angka"
        return hx_render("admin/siswa.jinja", push_url="admin.siswa", **notif)
    if jumlah_siswa <= 0:
        notif["error"] = "Jumlah siswa (sel B2) harus lebih dari 0"
        return hx_render("admin/siswa.jinja", push_url="admin.siswa", **notif)

    header_row = 4
    expected_headers = ["NIS", "Nama", "Password", "Rombel", "Catatan Admin"]
    for col_idx, header in enumerate(expected_headers, start=1):
        cell_val = str(ws.cell(row=header_row, column=col_idx).value or "").strip()
        if cell_val != header:
            notif["error"] = (
                "Header kolom {} tidak valid: '{}' (seharusnya '{}')".format(
                    col_idx, cell_val, header
                )
            )
            return hx_render("admin/siswa.jinja", push_url="admin.siswa", **notif)

    data_start = header_row + 1
    rows = []
    validation_errors = []

    for i in range(jumlah_siswa):
        r = data_start + i
        nis = ws.cell(row=r, column=1).value
        nama = ws.cell(row=r, column=2).value
        password = ws.cell(row=r, column=3).value
        rombel_id_raw = ws.cell(row=r, column=4).value
        catatan = ws.cell(row=r, column=5).value

        nis = str(nis).strip() if nis is not None else ""
        nama = str(nama).strip() if nama is not None else ""
        password = str(password).strip() if password is not None else ""
        catatan = str(catatan).strip() if catatan is not None else ""

        try:
            rombel_id = int(rombel_id_raw) if rombel_id_raw is not None else None
        except (ValueError, TypeError):
            validation_errors.append(
                "Baris {}: Rombel id '{}' tidak valid".format(r, rombel_id_raw)
            )
            rombel_id = None

        if not nis:
            validation_errors.append("Baris {}: NIS wajib diisi".format(r))
        if not nama:
            validation_errors.append("Baris {}: Nama wajib diisi".format(r))
        if rombel_id is not None and rombel_id not in valid_cg_ids:
            validation_errors.append(
                "Baris {}: Rombel id {} tidak valid".format(r, rombel_id)
            )

        rows.append(
            {
                "row_num": r,
                "nis": nis,
                "nama": nama,
                "password": password or nis,
                "rombel_id": rombel_id,
                "catatan": catatan,
            }
        )

    if validation_errors:
        notif["error"] = "Validasi gagal:<br>- " + "<br>- ".join(validation_errors)
        return hx_render("admin/siswa.jinja", push_url="admin.siswa", **notif)

    imported = 0
    updated = 0
    skipped = 0

    for row_data in rows:
        existing = Student.query.filter_by(student_id=row_data["nis"]).first()

        if existing:
            if mode == "update":
                existing.name = sanitize(row_data["nama"])
                if row_data["password"]:
                    existing.password = generate_password_hash(
                        row_data["password"], method="pbkdf2:sha256", salt_length=16
                    )
                existing.class_group_id = row_data["rombel_id"]
                existing.admin_note = sanitize(row_data["catatan"]) or None
                existing.is_deleted = False
                updated += 1
            else:
                skipped += 1
        else:
            student = Student(
                student_id=sanitize(row_data["nis"]),
                name=sanitize(row_data["nama"]),
                password=generate_password_hash(
                    row_data["password"], method="pbkdf2:sha256", salt_length=16
                ),
                class_group_id=row_data["rombel_id"],
                admin_note=sanitize(row_data["catatan"]) or None,
            )
            db.session.add(student)
            imported += 1

    db.session.commit()

    parts = []
    if imported:
        parts.append("{} siswa ditambahkan".format(imported))
    if updated:
        parts.append("{} siswa diperbarui".format(updated))
    if skipped:
        parts.append("{} siswa dilewati (NIS sudah ada)".format(skipped))
    notif["success"] = "Import selesai: {}".format(", ".join(parts))

    return hx_render("admin/siswa.jinja", push_url="admin.siswa", **notif)


# ---- Guru (Admin) CRUD ----


@bp.route("/guru")
@superadmin_required
def guru():
    return hx_render("admin/guru.jinja")


@bp.route("/guru/data")
@superadmin_required
def guru_data():
    admins = Teacher.query.order_by(Teacher.id).all()
    data = []
    for i, a in enumerate(admins, 1):
        role = "Superadmin" if a.is_superadmin else "Admin"
        data.append(
            {
                "no": i,
                "username": a.username,
                "name": a.name or "-",
                "contact_person": a.contact_person or "-",
                "role": role,
                "actions": (
                    '<a class="btn btn-sm btn-warning" '
                    f'onclick="edit_guru({a.id})">'
                    '<i class="bi bi-pencil"></i> Edit</a> '
                    '<button type="button" class="btn btn-sm btn-danger" '
                    f"onclick=\"hapus_guru({a.id}, '{sanitize(a.username)}')\">"
                    '<i class="bi bi-trash"></i> Hapus</button>'
                ),
            }
        )
    return jsonify(data=data)


@bp.route("/guru/tambah", methods=["GET", "POST"])
@superadmin_required
def guru_tambah():
    form = GuruForm()
    if request.method == "GET":
        return hx_render("admin/guru_form.jinja", guru=None, form=form)

    if not form.validate_on_submit():
        return hx_render("admin/guru_form.jinja", guru=None, form=form)

    notif = {}
    existing = Teacher.query.filter_by(username=form.username.data).first()
    if existing:
        notif["error"] = "Username sudah digunakan"
        return hx_render("admin/guru_form.jinja", guru=None, form=form, **notif)

    if not form.password.data:
        form.password.errors.append("Password wajib diisi")
        return hx_render("admin/guru_form.jinja", guru=None, form=form)

    admin = Teacher(
        username=sanitize(form.username.data),
        name=sanitize(form.name.data),
        contact_person=sanitize(form.contact_person.data) or None,
        password=generate_password_hash(
            form.password.data, method="pbkdf2:sha256", salt_length=16
        ),
    )
    db.session.add(admin)
    db.session.commit()
    notif["success"] = "Guru berhasil ditambahkan"
    return hx_render("admin/guru.jinja", push_url="admin.guru", **notif)


@bp.route("/guru/edit/<int:id>", methods=["GET", "POST"])
@superadmin_required
def guru_edit(id):
    admin = Teacher.query.get_or_404(id)
    form = GuruForm(obj=admin)
    if request.method == "GET":
        return hx_render("admin/guru_form.jinja", guru=admin, form=form)

    if not form.validate_on_submit():
        return hx_render("admin/guru_form.jinja", guru=admin, form=form)

    notif = {}
    existing = Teacher.query.filter(
        Teacher.username == form.username.data,
        Teacher.id != id,
    ).first()
    if existing:
        notif["error"] = "Username sudah digunakan guru lain"
        return hx_render("admin/guru_form.jinja", guru=admin, form=form, **notif)

    admin.username = sanitize(form.username.data)
    admin.name = sanitize(form.name.data)
    admin.contact_person = sanitize(form.contact_person.data) or None
    if form.password.data:
        admin.password = generate_password_hash(
            form.password.data, method="pbkdf2:sha256", salt_length=16
        )
    db.session.commit()
    notif["success"] = "Guru berhasil diperbarui"
    return hx_render("admin/guru.jinja", push_url="admin.guru", **notif)


@bp.route("/guru/hapus", methods=["POST"])
@superadmin_required
def guru_hapus():
    id = request.form.get("id", type=int)
    admin = Teacher.query.get_or_404(id)
    notif = {}
    if admin.username == session["admin_name"]:
        notif["error"] = "Tidak dapat menghapus akun yang sedang digunakan"
    else:
        db.session.delete(admin)
        db.session.commit()
        notif["success"] = "Guru berhasil dihapus"
    return hx_render("admin/guru.jinja", push_url="admin.guru", **notif)


# ---- Kategori (Category) CRUD ----


def _populate_kategori_teacher_choices(form):
    form.teachers.choices = [
        (t.id, t.name or t.username)
        for t in Teacher.query.order_by(Teacher.username).all()
    ]


@bp.route("/kategori")
@admin_required
def kategori():
    return hx_render("admin/kategori.jinja")


@bp.route("/kategori/data")
@admin_required
def kategori_data():
    from sqlalchemy.orm import joinedload

    categories = (
        Category.query.options(joinedload(Category.teacher_links))
        .order_by(Category.id)
        .all()
    )
    teacher_ids = {link.teacher_id for cat in categories for link in cat.teacher_links}
    teacher_map = {}
    if teacher_ids:
        for t in Teacher.query.filter(Teacher.id.in_(teacher_ids)).all():
            teacher_map[t.id] = t.name or t.username

    is_superadmin = session.get("is_superadmin", False)
    data = []
    for i, cat in enumerate(categories, 1):
        teacher_names = [
            teacher_map.get(link.teacher_id, "[tidak dikenal]")
            for link in cat.teacher_links
        ]
        data.append(
            {
                "no": i,
                "name": cat.name,
                "teachers": ", ".join(teacher_names) if teacher_names else "-",
                "actions": (
                    '<a class="btn btn-sm btn-warning" '
                    f'onclick="edit_kategori({cat.id})">'
                    '<i class="bi bi-pencil"></i> Edit</a> '
                    '<button type="button" class="btn btn-sm btn-danger" '
                    f"onclick=\"hapus_kategori({cat.id}, '{sanitize(cat.name)}')\">"
                    '<i class="bi bi-trash"></i> Hapus</button>'
                )
                if is_superadmin
                else "-",
            }
        )
    return jsonify(data=data)


@bp.route("/kategori/tambah", methods=["GET", "POST"])
@superadmin_required
def kategori_tambah():
    form = KategoriForm()
    _populate_kategori_teacher_choices(form)
    if request.method == "GET":
        return hx_render("admin/kategori_form.jinja", category=None, form=form)

    if not form.validate_on_submit():
        return hx_render("admin/kategori_form.jinja", category=None, form=form)

    notif = {}
    existing = Category.query.filter_by(name=form.name.data).first()
    if existing:
        notif["error"] = "Nama kategori sudah digunakan"
        return hx_render("admin/kategori_form.jinja", category=None, form=form, **notif)

    category = Category(name=sanitize(form.name.data))
    db.session.add(category)
    db.session.flush()
    for teacher_id in form.teachers.data:
        db.session.add(CategoryTeacher(category_id=category.id, teacher_id=teacher_id))
    db.session.commit()
    notif["success"] = "Kategori berhasil ditambahkan"
    return hx_render("admin/kategori.jinja", push_url="admin.kategori", **notif)


@bp.route("/kategori/edit/<int:id>", methods=["GET", "POST"])
@superadmin_required
def kategori_edit(id):
    category = Category.query.get_or_404(id)
    form = KategoriForm(obj=category)
    _populate_kategori_teacher_choices(form)
    if request.method == "GET":
        form.teachers.data = [link.teacher_id for link in category.teacher_links]
        return hx_render("admin/kategori_form.jinja", category=category, form=form)

    if not form.validate_on_submit():
        return hx_render("admin/kategori_form.jinja", category=category, form=form)

    notif = {}
    existing = Category.query.filter(
        Category.name == form.name.data,
        Category.id != id,
    ).first()
    if existing:
        notif["error"] = "Nama kategori sudah digunakan"
        return hx_render(
            "admin/kategori_form.jinja", category=category, form=form, **notif
        )

    category.name = sanitize(form.name.data)
    CategoryTeacher.query.filter_by(category_id=id).delete()
    for teacher_id in form.teachers.data:
        db.session.add(CategoryTeacher(category_id=id, teacher_id=teacher_id))
    db.session.commit()
    notif["success"] = "Kategori berhasil diperbarui"
    return hx_render("admin/kategori.jinja", push_url="admin.kategori", **notif)


@bp.route("/kategori/hapus", methods=["POST"])
@superadmin_required
def kategori_hapus():
    id = request.form.get("id", type=int)
    category = Category.query.get_or_404(id)
    db.session.delete(category)
    db.session.commit()
    notif = {"success": "Kategori berhasil dihapus"}
    return hx_render("admin/kategori.jinja", push_url="admin.kategori", **notif)


# ---- Permintaan Peminjaman ----


def _get_current_teacher():
    return Teacher.query.filter_by(username=session["admin_name"]).first()


def _get_teacher_category_ids(teacher_id):
    return [
        ct.category_id
        for ct in CategoryTeacher.query.filter_by(teacher_id=teacher_id).all()
    ]


def _teacher_can_review(teacher_id, category_id):
    return (
        CategoryTeacher.query.filter_by(
            teacher_id=teacher_id, category_id=category_id
        ).first()
        is not None
    )


@bp.route("/permintaan/data")
@admin_required
def permintaan_data():
    from sqlalchemy.orm import joinedload

    teacher = _get_current_teacher()
    is_superadmin = session.get("is_superadmin", False)
    teacher_category_ids = _get_teacher_category_ids(teacher.id)

    query = BorrowingRequest.query.options(
        joinedload(BorrowingRequest.student).joinedload(Student.class_group),
        joinedload(BorrowingRequest.category),
    )

    if not is_superadmin:
        query = query.filter(BorrowingRequest.category_id.in_(teacher_category_ids))

    requests = query.order_by(BorrowingRequest.date.desc()).all()

    status_badges = {
        "pending": '<span class="badge bg-warning text-dark">Pending</span>',
        "accepted": '<span class="badge bg-success">Diterima</span>',
        "rejected": '<span class="badge bg-danger">Ditolak</span>',
    }

    data = []
    for i, req in enumerate(requests, 1):
        student = req.student
        cat_group = (
            student.class_group.display_name if student and student.class_group else "-"
        )
        student_name = student.name if student else "-"
        student_nis = student.student_id if student else "-"
        category_name = req.category.name if req.category else "-"

        detail_btn = (
            '<a class="btn btn-sm btn-info text-white" '
            f'onclick="detail_permintaan({req.id})">'
            '<i class="bi bi-eye"></i> Detail</a>'
        )

        data.append(
            {
                "no": i,
                "date": req.date.strftime("%d/%m/%Y") if req.date else "-",
                "student_nis": student_nis,
                "student_name": student_name,
                "class_group": cat_group,
                "category": category_name,
                "status": status_badges.get(req.status, req.status),
                "actions": detail_btn,
            }
        )
    return jsonify(data=data)


@bp.route("/permintaan/<int:id>")
@admin_required
def permintaan_detail(id):
    from sqlalchemy.orm import joinedload

    teacher = _get_current_teacher()
    req = BorrowingRequest.query.options(
        joinedload(BorrowingRequest.student).joinedload(Student.class_group),
        joinedload(BorrowingRequest.category),
        joinedload(BorrowingRequest.reviewer),
    ).get_or_404(id)
    can_review = _teacher_can_review(teacher.id, req.category_id)
    category_teachers = (
        Teacher.query.join(CategoryTeacher, CategoryTeacher.teacher_id == Teacher.id)
        .filter(CategoryTeacher.category_id == req.category_id)
        .all()
    )

    return hx_render(
        "admin/permintaan_detail.jinja",
        req=req,
        can_review=can_review,
        category_teachers=category_teachers,
    )


@bp.route("/permintaan/terima/<int:id>", methods=["POST"])
@admin_required
def permintaan_terima(id):
    teacher = _get_current_teacher()
    req = BorrowingRequest.query.get_or_404(id)

    if not _teacher_can_review(teacher.id, req.category_id):
        return hx_render(
            "admin/permintaan_detail.jinja",
            req=req,
            can_review=False,
            error="Anda tidak berwenang meninjau permintaan ini",
            push_url=url_for("admin.permintaan_detail", id=id),
        )

    if req.status != "pending":
        return hx_render(
            "admin/permintaan_detail.jinja",
            req=req,
            can_review=True,
            error="Permintaan sudah ditinjau",
            push_url=url_for("admin.permintaan_detail", id=id),
        )

    req.status = "accepted"
    req.reviewed_by = teacher.id
    req.teacher_note = sanitize(request.form.get("teacher_note")) or None
    req.reviewed_at = datetime.now(timezone.utc)
    db.session.commit()

    return hx_render(
        "admin/permintaan_detail.jinja",
        req=req,
        can_review=True,
        success="Permintaan berhasil diterima",
        push_url=url_for("admin.permintaan_detail", id=id),
    )


@bp.route("/permintaan/tolak/<int:id>", methods=["POST"])
@admin_required
def permintaan_tolak(id):
    teacher = _get_current_teacher()
    req = BorrowingRequest.query.get_or_404(id)

    if not _teacher_can_review(teacher.id, req.category_id):
        return hx_render(
            "admin/permintaan_detail.jinja",
            req=req,
            can_review=False,
            error="Anda tidak berwenang meninjau permintaan ini",
            push_url=url_for("admin.permintaan_detail", id=id),
        )

    if req.status != "pending":
        return hx_render(
            "admin/permintaan_detail.jinja",
            req=req,
            can_review=True,
            error="Permintaan sudah ditinjau",
            push_url=url_for("admin.permintaan_detail", id=id),
        )

    req.status = "rejected"
    req.reviewed_by = teacher.id
    req.teacher_note = sanitize(request.form.get("teacher_note")) or None
    req.reviewed_at = datetime.now(timezone.utc)
    db.session.commit()

    return hx_render(
        "admin/permintaan_detail.jinja",
        req=req,
        can_review=True,
        success="Permintaan berhasil ditolak",
        push_url=url_for("admin.permintaan_detail", id=id),
    )


@bp.route("/permintaan/batalkan/<int:id>", methods=["POST"])
@admin_required
def permintaan_batalkan(id):
    teacher = _get_current_teacher()
    req = BorrowingRequest.query.get_or_404(id)

    if not _teacher_can_review(teacher.id, req.category_id):
        return hx_render(
            "admin/permintaan_detail.jinja",
            req=req,
            can_review=False,
            error="Anda tidak berwenang meninjau permintaan ini",
            push_url=url_for("admin.permintaan_detail", id=id),
        )

    now = datetime.now(WIB)
    if req.date < now.date() or (req.date == now.date() and now.hour >= 17):
        return hx_render(
            "admin/permintaan_detail.jinja",
            req=req,
            can_review=True,
            error="Pembatalan tidak dapat dilakukan setelah pukul 17:00 pada hari peminjaman atau untuk tanggal yang sudah lewat",
            push_url=url_for("admin.permintaan_detail", id=id),
        )

    if req.status not in ("accepted", "rejected"):
        return hx_render(
            "admin/permintaan_detail.jinja",
            req=req,
            can_review=True,
            error="Permintaan belum ditinjau",
            push_url=url_for("admin.permintaan_detail", id=id),
        )

    req.status = "pending"
    req.reviewed_by = None
    req.teacher_note = None
    req.reviewed_at = None
    db.session.commit()

    return hx_render(
        "admin/permintaan_detail.jinja",
        req=req,
        can_review=True,
        success="Keputusan berhasil dibatalkan",
        push_url=url_for("admin.permintaan_detail", id=id),
    )
