from datetime import datetime, timedelta, timezone

from flask import Blueprint, jsonify, request, send_file, session

from .db import db
from .helper import WIB, admin_required, hx_render
from .models import BorrowingRequest, Student, Teacher

bp = Blueprint("supervisor", __name__, url_prefix="/supervisor")


def _get_current_teacher():
    return Teacher.query.filter_by(username=session.get("admin_name")).first()


def _parse_tanggal(tanggal_str, default):
    if not tanggal_str:
        return default
    try:
        return datetime.strptime(tanggal_str, "%Y-%m-%d").date()
    except ValueError:
        return default


@bp.route("/monitor")
@admin_required
def monitor():
    today = datetime.now(WIB).date()
    tanggal = _parse_tanggal(request.args.get("tanggal"), today)

    return hx_render(
        "supervisor/monitor.jinja",
        tanggal=tanggal,
        min_date=today - timedelta(days=365),
        max_date=today + timedelta(days=365),
    )


@bp.route("/monitor/data")
@admin_required
def monitor_data():
    from sqlalchemy.orm import joinedload

    today = datetime.now(WIB).date()
    tanggal = _parse_tanggal(request.args.get("tanggal"), today)

    requests = (
        BorrowingRequest.query.filter_by(status="accepted", date=tanggal)
        .options(
            joinedload(BorrowingRequest.student).joinedload(
                Student.class_group
            ),
            joinedload(BorrowingRequest.category),
            joinedload(BorrowingRequest.reviewer),
            joinedload(BorrowingRequest.confirmer),
        )
        .order_by(BorrowingRequest.id)
        .all()
    )

    data = []
    for i, req in enumerate(requests, 1):
        student = req.student
        data.append(
            {
                "no": i,
                "request_id": req.id,
                "student_nis": student.student_id if student else "-",
                "student_name": student.name if student else "-",
                "class_group": (
                    student.class_group.display_name
                    if student and student.class_group
                    else "-"
                ),
                "category": req.category.name if req.category else "-",
                "student_note": req.student_note or "-",
                "reviewer": req.reviewer.name if req.reviewer else "-",
                "confirmation": req.confirmation,
                "confirmed_by": (req.confirmer.name if req.confirmer else None),
                "confirmed_at": (
                    req.confirmed_at.strftime("%d/%m/%Y %H:%M")
                    if req.confirmed_at
                    else None
                ),
            }
        )
    return jsonify(data=data)


@bp.route("/monitor/export")
@admin_required
def monitor_export():
    from io import BytesIO

    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from sqlalchemy.orm import joinedload

    today = datetime.now(WIB).date()
    tanggal_mulai = _parse_tanggal(
        request.args.get("tanggal_mulai"), today - timedelta(days=30)
    )
    tanggal_akhir = _parse_tanggal(request.args.get("tanggal_akhir"), today)

    if tanggal_mulai > tanggal_akhir:
        tanggal_mulai, tanggal_akhir = tanggal_akhir, tanggal_mulai

    hanya_digunakan = request.args.get("hanya_digunakan") == "1"

    query = (
        BorrowingRequest.query.filter(
            BorrowingRequest.status == "accepted",
            BorrowingRequest.date.between(tanggal_mulai, tanggal_akhir),
        )
        .options(
            joinedload(BorrowingRequest.student).joinedload(
                Student.class_group
            ),
            joinedload(BorrowingRequest.category),
            joinedload(BorrowingRequest.reviewer),
            joinedload(BorrowingRequest.confirmer),
        )
        .order_by(BorrowingRequest.date, BorrowingRequest.id)
    )
    if hanya_digunakan:
        query = query.filter(BorrowingRequest.confirmation == "used")

    requests = query.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Peminjaman"

    judul_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")

    ws["A1"] = "Laporan Peminjaman Laptop"
    ws["A1"].font = judul_font
    ws["A2"] = "Periode: {} s/d {}".format(
        tanggal_mulai.strftime("%d/%m/%Y"),
        tanggal_akhir.strftime("%d/%m/%Y"),
    )
    ws["A3"] = "Filter: {}".format(
        "Hanya yang digunakan" if hanya_digunakan else "Semua permintaan"
    )

    headers = [
        "No",
        "Tanggal",
        "NIS",
        "Nama Siswa",
        "Rombel",
        "Kategori",
        "Catatan Siswa",
        "Penanggung Jawab",
        "Konfirmasi",
        "Dikonfirmasi Oleh",
        "Waktu Konfirmasi",
    ]
    header_row = 5
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    konfirmasi_label = {
        "used": "Digunakan",
        "not_used": "Tidak Digunakan",
    }
    for i, req in enumerate(requests, 1):
        student = req.student
        row = header_row + i
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=req.date.strftime("%d/%m/%Y"))
        ws.cell(row=row, column=3, value=student.student_id if student else "-")
        ws.cell(row=row, column=4, value=student.name if student else "-")
        ws.cell(
            row=row,
            column=5,
            value=(
                student.class_group.display_name
                if student and student.class_group
                else "-"
            ),
        )
        ws.cell(
            row=row, column=6, value=req.category.name if req.category else "-"
        )
        ws.cell(row=row, column=7, value=req.student_note or "-")
        ws.cell(
            row=row, column=8, value=req.reviewer.name if req.reviewer else "-"
        )
        ws.cell(
            row=row,
            column=9,
            value=konfirmasi_label.get(req.confirmation, "Belum Dikonfirmasi"),
        )
        ws.cell(
            row=row,
            column=10,
            value=req.confirmer.name if req.confirmer else "-",
        )
        ws.cell(
            row=row,
            column=11,
            value=(
                req.confirmed_at.strftime("%d/%m/%Y %H:%M")
                if req.confirmed_at
                else "-"
            ),
        )

    col_widths = [5, 12, 12, 25, 15, 15, 25, 20, 18, 20, 20]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    download_name = "laporan_peminjaman_{}_{}.xlsx".format(
        tanggal_mulai.strftime("%Y%m%d"),
        tanggal_akhir.strftime("%Y%m%d"),
    )
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=download_name,
    )


@bp.route("/monitor/konfirmasi/<int:id>", methods=["POST"])
@admin_required
def monitor_konfirmasi(id):
    teacher = _get_current_teacher()
    if not teacher:
        return jsonify(success=False, message="Sesi tidak valid"), 403

    req = db.get_or_404(BorrowingRequest, id)

    if req.status != "accepted":
        return (
            jsonify(
                success=False,
                message="Hanya permintaan yang diterima dapat dikonfirmasi",
            ),
            400,
        )

    confirmation = request.form.get("confirmation")
    if confirmation not in ("used", "not_used"):
        return (
            jsonify(success=False, message="Nilai konfirmasi tidak valid"),
            400,
        )

    req.confirmation = confirmation
    req.confirmed_by = teacher.id
    req.confirmed_at = datetime.now(timezone.utc)
    db.session.commit()

    label = "digunakan" if confirmation == "used" else "tidak digunakan"
    return jsonify(
        success=True, message="Konfirmasi berhasil: {}".format(label)
    )


@bp.route("/monitor/batalkan_konfirmasi/<int:id>", methods=["POST"])
@admin_required
def monitor_batalkan_konfirmasi(id):
    teacher = _get_current_teacher()
    if not teacher:
        return jsonify(success=False, message="Sesi tidak valid"), 403

    req = db.get_or_404(BorrowingRequest, id)

    if req.status != "accepted":
        return (
            jsonify(
                success=False,
                message="Hanya permintaan yang diterima dapat dikonfirmasi",
            ),
            400,
        )

    if not req.confirmation:
        return (
            jsonify(success=False, message="Permintaan belum dikonfirmasi"),
            400,
        )

    req.confirmation = None
    req.confirmed_by = None
    req.confirmed_at = None
    db.session.commit()

    return jsonify(success=True, message="Konfirmasi berhasil dibatalkan")
