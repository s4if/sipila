from app import db


def test_admin_dashboard_redirects_anonymous(client):
    response = client.get("/admin/")
    assert response.status_code == 302
    assert "login" in response.location


def test_admin_dashboard_returns_200_when_logged_in(logged_in_client):
    response = logged_in_client.get("/admin/")
    assert response.status_code == 200


def test_admin_password_page_redirects_anonymous(client):
    response = client.get("/admin/ganti_password")
    assert response.status_code == 302
    assert "login" in response.location


def test_admin_password_page_get(logged_in_client):
    response = logged_in_client.get("/admin/ganti_password")
    assert response.status_code == 200
    assert b"password" in response.data.lower() or b"ganti" in response.data


def test_change_password_success(logged_in_client):
    response = logged_in_client.post(
        "/admin/ganti_password",
        data={
            "current_password": "secret",
            "new_password": "newsecret123",
            "confirm_password": "newsecret123",
        },
    )
    assert response.status_code == 200
    assert b"berhasil" in response.data.lower()


def test_change_password_mismatch(logged_in_client):
    response = logged_in_client.post(
        "/admin/ganti_password",
        data={
            "current_password": "secret",
            "new_password": "newpass1",
            "confirm_password": "newpass2",
        },
    )
    assert response.status_code == 200
    assert b"tidak sesuai" in response.data


def test_change_password_wrong_current(logged_in_client):
    response = logged_in_client.post(
        "/admin/ganti_password",
        data={
            "current_password": "wrongpassword",
            "new_password": "newsecret123",
            "confirm_password": "newsecret123",
        },
    )
    assert response.status_code == 200
    assert b"tidak sesuai" in response.data or b"salah" in response.data


# ---- Guru CRUD tests ----


def test_guru_list_redirects_anonymous(client):
    response = client.get("/admin/guru")
    assert response.status_code == 302
    assert "login" in response.location


def test_guru_list_returns_200(logged_in_client):
    response = logged_in_client.get("/admin/guru")
    assert response.status_code == 200
    assert b"Data Guru" in response.data


def test_guru_data_returns_json(logged_in_client, admin_user):
    response = logged_in_client.get("/admin/guru/data")
    assert response.status_code == 200
    data = response.get_json()
    assert "data" in data
    assert len(data["data"]) >= 1
    assert data["data"][0]["username"] == "admin"


def test_guru_tambah_page_get(logged_in_client):
    response = logged_in_client.get("/admin/guru/tambah")
    assert response.status_code == 200
    assert b"Tambah Guru" in response.data


def test_guru_tambah_success(logged_in_client, app):
    response = logged_in_client.post(
        "/admin/guru/tambah",
        data={
            "username": "guru1",
            "name": "Guru Satu",
            "contact_person": "081234567890",
            "password": "password123",
        },
    )
    assert response.status_code == 200
    assert b"berhasil ditambahkan" in response.data

    with app.app_context():
        from app.models import Teacher

        guru = Teacher.query.filter_by(username="guru1").first()
        assert guru is not None
        assert guru.name == "Guru Satu"
        assert guru.is_superadmin is False


def test_guru_tambah_duplicate_username(logged_in_client):
    response = logged_in_client.post(
        "/admin/guru/tambah",
        data={
            "username": "admin",
            "name": "Duplikat",
            "password": "password123",
        },
    )
    assert response.status_code == 200
    assert b"sudah digunakan" in response.data


def test_guru_tambah_without_password(logged_in_client):
    response = logged_in_client.post(
        "/admin/guru/tambah",
        data={
            "username": "guru2",
            "name": "Guru Dua",
            "password": "",
        },
    )
    assert response.status_code == 200
    assert b"wajib diisi" in response.data


def test_guru_edit_page_get(logged_in_client, admin_user):
    response = logged_in_client.get(f"/admin/guru/edit/{admin_user.id}")
    assert response.status_code == 200
    assert b"Edit Guru" in response.data


def test_guru_edit_success(logged_in_client, app, admin_user):
    response = logged_in_client.post(
        f"/admin/guru/edit/{admin_user.id}",
        data={
            "username": "admin",
            "name": "Admin Baru",
            "contact_person": "089999",
            "password": "",
        },
    )
    assert response.status_code == 200
    assert b"berhasil diperbarui" in response.data

    with app.app_context():
        from app.models import Teacher

        guru = db.session.get(Teacher, admin_user.id)
        assert guru.name == "Admin Baru"


def test_guru_edit_duplicate_username(logged_in_client, app):
    from werkzeug.security import generate_password_hash

    from app import db
    from app.models import Teacher

    with app.app_context():
        guru2 = Teacher(
            username="guru2",
            password=generate_password_hash("pass"),
        )
        db.session.add(guru2)
        db.session.commit()
        guru2_id = guru2.id

    response = logged_in_client.post(
        f"/admin/guru/edit/{guru2_id}",
        data={
            "username": "admin",
            "name": "Guru Dua",
            "password": "",
        },
    )
    assert response.status_code == 200
    assert b"sudah digunakan" in response.data


def test_guru_edit_with_new_password(logged_in_client, app, admin_user):
    response = logged_in_client.post(
        f"/admin/guru/edit/{admin_user.id}",
        data={
            "username": "admin",
            "name": "Admin",
            "password": "newpassword456",
        },
    )
    assert response.status_code == 200
    assert b"berhasil diperbarui" in response.data

    with app.app_context():
        from werkzeug.security import check_password_hash

        from app.models import Teacher

        guru = db.session.get(Teacher, admin_user.id)
        assert check_password_hash(guru.password, "newpassword456")


def test_guru_hapus_success(logged_in_client, app):
    from werkzeug.security import generate_password_hash

    from app import db
    from app.models import Teacher

    with app.app_context():
        guru = Teacher(
            username="hapus_guru",
            password=generate_password_hash("pass"),
        )
        db.session.add(guru)
        db.session.commit()
        guru_id = guru.id

    response = logged_in_client.post("/admin/guru/hapus", data={"id": guru_id})
    assert response.status_code == 200
    assert b"berhasil dihapus" in response.data


def test_guru_hapus_self_blocked(logged_in_client, admin_user):
    response = logged_in_client.post(
        "/admin/guru/hapus",
        data={
            "id": admin_user.id,
        },
    )
    assert response.status_code == 200
    assert b"tidak dapat menghapus" in response.data.lower()


def test_guru_hapus_404(logged_in_client):
    response = logged_in_client.post("/admin/guru/hapus", data={"id": 9999})
    assert response.status_code == 404


# ---- Siswa Import tests ----


def _create_class_group(name="1", grade_level="X", major="TJKT"):
    from app import db
    from app.models import ClassGroup

    cg = ClassGroup(name=name, grade_level=grade_level, major=major)
    db.session.add(cg)
    db.session.commit()
    return cg.id


def _build_xlsx(mode="skip", jumlah_siswa=0, rombel_info=None, rows=None):
    from io import BytesIO

    import openpyxl

    if rombel_info is None:
        rombel_info = []
    if rows is None:
        rows = []

    wb = openpyxl.Workbook()
    ws = wb.active

    ws["B1"] = mode
    ws["B2"] = jumlah_siswa
    ws["H1"] = "Nama Rombel"
    ws["I1"] = "id"

    for i, (rname, rid) in enumerate(rombel_info, start=2):
        ws.cell(row=i, column=8, value=rname)
        ws.cell(row=i, column=9, value=rid)

    header_row = 4
    ws.cell(row=header_row, column=1, value="NIS")
    ws.cell(row=header_row, column=2, value="Nama")
    ws.cell(row=header_row, column=3, value="Password")
    ws.cell(row=header_row, column=4, value="Rombel")
    ws.cell(row=header_row, column=5, value="Catatan Admin")

    for i, row_data in enumerate(rows):
        r = header_row + 1 + i
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=r, column=col_idx, value=val)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_siswa_template_download(logged_in_client):
    response = logged_in_client.get("/admin/siswa/template")
    assert response.status_code == 200
    assert response.content_type == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert b"PK" in response.data


def test_siswa_import_no_file(logged_in_client):
    response = logged_in_client.post("/admin/siswa/import")
    assert response.status_code == 200
    assert b"Tidak ada file" in response.data


def test_siswa_import_wrong_extension(logged_in_client):
    from io import BytesIO

    buf = BytesIO(b"not an xlsx")
    response = logged_in_client.post(
        "/admin/siswa/import",
        data={"file": (buf, "data.csv")},
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    assert b"harus berformat" in response.data


def test_siswa_import_skip_mode(logged_in_client, app):
    from werkzeug.security import generate_password_hash

    from app import db
    from app.models import Student

    with app.app_context():
        cg_id = _create_class_group()
        existing = Student(
            student_id="1001",
            name="Lama",
            password=generate_password_hash("old"),
            class_group_id=cg_id,
        )
        db.session.add(existing)
        db.session.commit()

        buf = _build_xlsx(
            mode="skip",
            jumlah_siswa=2,
            rombel_info=[("X TJKT 1", cg_id)],
            rows=[
                ("1001", "Baru", "pass123", cg_id, ""),
                ("1002", "Ahmad", "", cg_id, ""),
            ],
        )

    response = logged_in_client.post(
        "/admin/siswa/import",
        data={"file": (buf, "import.xlsx")},
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    assert b"ditambahkan" in response.data
    assert b"dilewati" in response.data

    with app.app_context():
        assert Student.query.filter_by(student_id="1001").first().name == "Lama"
        assert Student.query.filter_by(student_id="1002").first() is not None


def test_siswa_import_update_mode(logged_in_client, app):
    from werkzeug.security import generate_password_hash

    from app import db
    from app.models import Student

    with app.app_context():
        cg_id = _create_class_group()
        existing = Student(
            student_id="1001",
            name="Lama",
            password=generate_password_hash("old"),
            class_group_id=cg_id,
        )
        db.session.add(existing)
        db.session.commit()

        buf = _build_xlsx(
            mode="update",
            jumlah_siswa=1,
            rombel_info=[("X TJKT 1", cg_id)],
            rows=[
                ("1001", "Baru", "newpass", cg_id, ""),
            ],
        )

    response = logged_in_client.post(
        "/admin/siswa/import",
        data={"file": (buf, "import.xlsx")},
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    assert b"diperbarui" in response.data

    with app.app_context():
        s = Student.query.filter_by(student_id="1001").first()
        assert s.name == "Baru"


def test_siswa_import_default_password_is_nis(logged_in_client, app):
    from werkzeug.security import check_password_hash

    from app.models import Student

    with app.app_context():
        cg_id = _create_class_group()

        buf = _build_xlsx(
            mode="skip",
            jumlah_siswa=1,
            rombel_info=[("X TJKT 1", cg_id)],
            rows=[
                ("2001", "Siti", "", cg_id, ""),
            ],
        )

    response = logged_in_client.post(
        "/admin/siswa/import",
        data={"file": (buf, "import.xlsx")},
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    assert b"ditambahkan" in response.data

    with app.app_context():
        s = Student.query.filter_by(student_id="2001").first()
        assert s is not None
        assert check_password_hash(s.password, "2001")


def test_siswa_import_abort_invalid_rombel(logged_in_client, app):

    with app.app_context():
        cg_id = _create_class_group()

        buf = _build_xlsx(
            mode="skip",
            jumlah_siswa=1,
            rombel_info=[("X TJKT 1", cg_id)],
            rows=[
                ("3001", "Budi", "pass", 9999, ""),
            ],
        )

    response = logged_in_client.post(
        "/admin/siswa/import",
        data={"file": (buf, "import.xlsx")},
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    assert b"tidak valid" in response.data


def test_siswa_import_abort_missing_nis(logged_in_client, app):

    with app.app_context():
        cg_id = _create_class_group()

        buf = _build_xlsx(
            mode="skip",
            jumlah_siswa=1,
            rombel_info=[("X TJKT 1", cg_id)],
            rows=[
                ("", "Budi", "pass", cg_id, ""),
            ],
        )

    response = logged_in_client.post(
        "/admin/siswa/import",
        data={"file": (buf, "import.xlsx")},
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    assert b"NIS wajib" in response.data


def test_siswa_import_abort_missing_nama(logged_in_client, app):

    with app.app_context():
        cg_id = _create_class_group()

        buf = _build_xlsx(
            mode="skip",
            jumlah_siswa=1,
            rombel_info=[("X TJKT 1", cg_id)],
            rows=[
                ("4001", "", "pass", cg_id, ""),
            ],
        )

    response = logged_in_client.post(
        "/admin/siswa/import",
        data={"file": (buf, "import.xlsx")},
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    assert b"Nama wajib" in response.data


def test_siswa_import_abort_invalid_rombel_in_reference(logged_in_client, app):

    with app.app_context():
        cg_id = _create_class_group()

        buf = _build_xlsx(
            mode="skip",
            jumlah_siswa=0,
            rombel_info=[("X TJKT 1", cg_id), ("Fake Class", 9999)],
        )

    response = logged_in_client.post(
        "/admin/siswa/import",
        data={"file": (buf, "import.xlsx")},
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    assert b"Rombel id tidak valid" in response.data


def test_siswa_import_zero_jumlah(logged_in_client, app):

    with app.app_context():
        buf = _build_xlsx(mode="skip", jumlah_siswa=0, rombel_info=[])

    response = logged_in_client.post(
        "/admin/siswa/import",
        data={"file": (buf, "import.xlsx")},
        content_type="multipart/form-data",
    )
    assert response.status_code == 200
    assert b"harus lebih dari 0" in response.data


# ---- Rombel (ClassGroup) CRUD tests ----


def test_rombel_list_redirects_anonymous(client):
    response = client.get("/admin/rombel")
    assert response.status_code == 302
    assert "login" in response.location


def test_rombel_list_returns_200(logged_in_client):
    response = logged_in_client.get("/admin/rombel")
    assert response.status_code == 200
    assert b"Data Rombel" in response.data


def test_rombel_data_returns_json(logged_in_client):
    response = logged_in_client.get("/admin/rombel/data")
    assert response.status_code == 200
    data = response.get_json()
    assert "data" in data


def test_rombel_data_includes_groups(logged_in_client, app):
    with app.app_context():
        _create_class_group(name="1", grade_level="X", major="TJKT")
        _create_class_group(name="2", grade_level="XI", major="TJKT")

    response = logged_in_client.get("/admin/rombel/data")
    assert response.status_code == 200
    data = response.get_json()
    assert len(data["data"]) == 2


def test_rombel_tambah_page_get(logged_in_client):
    response = logged_in_client.get("/admin/rombel/tambah")
    assert response.status_code == 200
    assert b"Tambah Rombel" in response.data


def test_rombel_tambah_success(logged_in_client, app):
    response = logged_in_client.post(
        "/admin/rombel/tambah",
        data={
            "grade_level": "X",
            "major": "TJKT",
            "name": "5",
            "homeroom_teacher_id": "",
        },
    )
    assert response.status_code == 200
    assert b"berhasil ditambahkan" in response.data

    with app.app_context():
        from app.models import ClassGroup

        cg = ClassGroup.query.filter_by(name="5", grade_level="X").first()
        assert cg is not None


def test_rombel_tambah_regular_admin_blocked(regular_admin_client):
    response = regular_admin_client.get("/admin/rombel/tambah")
    assert response.status_code == 302
    assert response.location == "/admin/"


def test_rombel_tambah_anonymous_redirect(client):
    response = client.get("/admin/rombel/tambah")
    assert response.status_code == 302
    assert "login" in response.location


def test_rombel_edit_page_get(logged_in_client, app):
    with app.app_context():
        cg_id = _create_class_group()

    response = logged_in_client.get("/admin/rombel/edit/{}".format(cg_id))
    assert response.status_code == 200
    assert b"Edit Rombel" in response.data


def test_rombel_edit_success(logged_in_client, app):
    with app.app_context():
        cg_id = _create_class_group(name="1", grade_level="X", major="TJKT")

    response = logged_in_client.post(
        "/admin/rombel/edit/{}".format(cg_id),
        data={
            "grade_level": "XI",
            "major": "",
            "name": "2",
            "homeroom_teacher_id": "",
        },
    )
    assert response.status_code == 200
    assert b"berhasil diperbarui" in response.data

    with app.app_context():
        from app.models import ClassGroup

        cg = db.session.get(ClassGroup, cg_id)
        assert cg.grade_level == "XI"
        assert cg.name == "2"
        assert cg.major is None


def test_rombel_edit_404(logged_in_client):
    response = logged_in_client.get("/admin/rombel/edit/9999")
    assert response.status_code == 404


def test_rombel_hapus_success(logged_in_client, app):
    with app.app_context():
        cg_id = _create_class_group()

    response = logged_in_client.post("/admin/rombel/hapus", data={"id": cg_id})
    assert response.status_code == 200
    assert b"berhasil dihapus" in response.data

    with app.app_context():
        from app.models import ClassGroup

        assert db.session.get(ClassGroup, cg_id) is None


def test_rombel_hapus_with_active_students_blocked(logged_in_client, app):
    from werkzeug.security import generate_password_hash

    with app.app_context():
        cg_id = _create_class_group()
        from app.models import Student

        s = Student(
            student_id="S700",
            name="Aktif",
            password=generate_password_hash("p"),
            class_group_id=cg_id,
        )
        db.session.add(s)
        db.session.commit()

    response = logged_in_client.post("/admin/rombel/hapus", data={"id": cg_id})
    assert response.status_code == 200
    assert b"tidak dapat dihapus" in response.data


def test_rombel_hapus_404(logged_in_client):
    response = logged_in_client.post("/admin/rombel/hapus", data={"id": 9999})
    assert response.status_code == 404


# ---- Siswa manual CRUD tests ----


def test_siswa_list_redirects_anonymous(client):
    response = client.get("/admin/siswa")
    assert response.status_code == 302
    assert "login" in response.location


def test_siswa_list_returns_200(logged_in_client):
    response = logged_in_client.get("/admin/siswa")
    assert response.status_code == 200
    assert b"Data Siswa" in response.data


def test_siswa_data_returns_json(logged_in_client):
    response = logged_in_client.get("/admin/siswa/data")
    assert response.status_code == 200
    data = response.get_json()
    assert "data" in data


def test_siswa_tambah_page_get(logged_in_client, app):
    with app.app_context():
        _create_class_group()

    response = logged_in_client.get("/admin/siswa/tambah")
    assert response.status_code == 200
    assert b"Tambah Siswa" in response.data


def test_siswa_tambah_success(logged_in_client, app):
    with app.app_context():
        cg_id = _create_class_group()

    response = logged_in_client.post(
        "/admin/siswa/tambah",
        data={
            "student_id": "S100",
            "name": "Budi",
            "password": "pass123",
            "class_group_id": str(cg_id),
            "admin_note": "",
        },
    )
    assert response.status_code == 200
    assert b"berhasil ditambahkan" in response.data

    with app.app_context():
        from app.models import Student

        s = Student.query.filter_by(student_id="S100").first()
        assert s is not None
        assert s.name == "Budi"


def test_siswa_tambah_duplicate_nis(logged_in_client, app):
    from werkzeug.security import generate_password_hash

    with app.app_context():
        cg_id = _create_class_group()
        from app.models import Student

        existing = Student(
            student_id="S100",
            name="Lama",
            password=generate_password_hash("p"),
            class_group_id=cg_id,
        )
        db.session.add(existing)
        db.session.commit()

    response = logged_in_client.post(
        "/admin/siswa/tambah",
        data={
            "student_id": "S100",
            "name": "Budi",
            "password": "pass123",
            "class_group_id": str(cg_id),
            "admin_note": "",
        },
    )
    assert response.status_code == 200
    assert b"NIS sudah terdaftar" in response.data


def test_siswa_tambah_no_password(logged_in_client, app):
    with app.app_context():
        cg_id = _create_class_group()

    response = logged_in_client.post(
        "/admin/siswa/tambah",
        data={
            "student_id": "S200",
            "name": "No Pass",
            "password": "",
            "class_group_id": str(cg_id),
            "admin_note": "",
        },
    )
    assert response.status_code == 200
    assert b"Password wajib diisi" in response.data


def test_siswa_tambah_regular_admin_blocked(regular_admin_client):
    response = regular_admin_client.get("/admin/siswa/tambah")
    assert response.status_code == 302
    assert response.location == "/admin/"


def test_siswa_edit_page_get(logged_in_client, app):
    from werkzeug.security import generate_password_hash

    with app.app_context():
        cg_id = _create_class_group()
        from app.models import Student

        s = Student(
            student_id="S300",
            name="Edit",
            password=generate_password_hash("p"),
            class_group_id=cg_id,
        )
        db.session.add(s)
        db.session.commit()
        s_id = s.id

    response = logged_in_client.get("/admin/siswa/edit/{}".format(s_id))
    assert response.status_code == 200
    assert b"Edit Siswa" in response.data


def test_siswa_edit_success(logged_in_client, app):
    from werkzeug.security import generate_password_hash

    with app.app_context():
        cg_id = _create_class_group()
        from app.models import Student

        s = Student(
            student_id="S400",
            name="Lama",
            password=generate_password_hash("p"),
            class_group_id=cg_id,
        )
        db.session.add(s)
        db.session.commit()
        s_id = s.id

    response = logged_in_client.post(
        "/admin/siswa/edit/{}".format(s_id),
        data={
            "student_id": "S401",
            "name": "Baru",
            "password": "",
            "class_group_id": str(cg_id),
            "admin_note": "catatan",
        },
    )
    assert response.status_code == 200
    assert b"berhasil diperbarui" in response.data

    with app.app_context():
        from app.models import Student

        s = db.session.get(Student, s_id)
        assert s.name == "Baru"
        assert s.student_id == "S401"


def test_siswa_edit_duplicate_nis(logged_in_client, app):
    from werkzeug.security import generate_password_hash

    with app.app_context():
        cg_id = _create_class_group()
        from app.models import Student

        s1 = Student(
            student_id="S500",
            name="A",
            password=generate_password_hash("p"),
            class_group_id=cg_id,
        )
        s2 = Student(
            student_id="S501",
            name="B",
            password=generate_password_hash("p"),
            class_group_id=cg_id,
        )
        db.session.add_all([s1, s2])
        db.session.commit()
        s2_id = s2.id

    response = logged_in_client.post(
        "/admin/siswa/edit/{}".format(s2_id),
        data={
            "student_id": "S500",
            "name": "B",
            "password": "",
            "class_group_id": str(cg_id),
            "admin_note": "",
        },
    )
    assert response.status_code == 200
    assert b"sudah digunakan" in response.data


def test_siswa_hapus_success(logged_in_client, app):
    from werkzeug.security import generate_password_hash

    with app.app_context():
        cg_id = _create_class_group()
        from app.models import Student

        s = Student(
            student_id="S600",
            name="Hapus",
            password=generate_password_hash("p"),
            class_group_id=cg_id,
        )
        db.session.add(s)
        db.session.commit()
        s_id = s.id

    response = logged_in_client.post("/admin/siswa/hapus", data={"id": s_id})
    assert response.status_code == 200
    assert b"berhasil dihapus" in response.data

    with app.app_context():
        from app.models import Student

        s = db.session.get(Student, s_id)
        assert s.is_deleted is True


def test_siswa_hapus_404(logged_in_client):
    response = logged_in_client.post("/admin/siswa/hapus", data={"id": 9999})
    assert response.status_code == 404


# ---- Siswa detail (riwayat peminjaman) tests ----


def _make_student_with_requests(app, dates_statuses):
    from datetime import datetime

    from werkzeug.security import generate_password_hash

    from app.models import (
        BorrowingRequest,
        Category,
        ClassGroup,
        Student,
    )

    with app.app_context():
        cg = ClassGroup(name="1", grade_level="X", major="TJKT")
        db.session.add(cg)
        db.session.flush()

        student = Student(
            student_id="S900",
            name="Siswa Riwayat",
            password=generate_password_hash("p"),
            class_group_id=cg.id,
        )
        db.session.add(student)
        db.session.flush()

        cat = Category(name="Laptop")
        db.session.add(cat)
        db.session.flush()

        for date_str, status in dates_statuses:
            req = BorrowingRequest(
                student_id=student.id,
                category_id=cat.id,
                date=datetime.strptime(date_str, "%Y-%m-%d").date(),
                status=status,
            )
            db.session.add(req)

        db.session.commit()
        return student.id


def test_siswa_detail_redirects_anonymous(client, siswa_user):
    response = client.get("/admin/siswa/{}".format(siswa_user.id))
    assert response.status_code == 302
    assert "login" in response.location


def test_siswa_detail_returns_200(logged_in_client, siswa_user):
    response = logged_in_client.get(
        "/admin/siswa/{}".format(siswa_user.id)
    )
    assert response.status_code == 200
    assert b"Riwayat Permintaan Peminjaman" in response.data
    assert siswa_user.name.encode() in response.data


def test_siswa_detail_404(logged_in_client):
    response = logged_in_client.get("/admin/siswa/9999")
    assert response.status_code == 404


def test_siswa_detail_data_returns_json(logged_in_client, app):
    student_id = _make_student_with_requests(
        app, [("2025-01-10", "accepted"), ("2025-02-10", "pending")]
    )
    response = logged_in_client.get(
        "/admin/siswa/{}/data".format(student_id)
    )
    assert response.status_code == 200
    data = response.get_json()
    assert "data" in data
    assert len(data["data"]) == 2


def test_siswa_detail_data_date_filter(logged_in_client, app):
    student_id = _make_student_with_requests(
        app,
        [("2025-01-10", "accepted"), ("2025-02-10", "pending"),
         ("2025-03-10", "rejected")],
    )
    response = logged_in_client.get(
        "/admin/siswa/{}/data".format(student_id),
        query_string={"start_date": "2025-02-01", "end_date": "2025-02-28"},
    )
    assert response.status_code == 200
    data = response.get_json()
    assert len(data["data"]) == 1
    assert data["data"][0]["date"] == "10/02/2025"


def test_siswa_detail_data_invalid_date_ignored(logged_in_client, app):
    student_id = _make_student_with_requests(
        app, [("2025-01-10", "accepted")]
    )
    response = logged_in_client.get(
        "/admin/siswa/{}/data".format(student_id),
        query_string={"start_date": "bukan-tanggal"},
    )
    assert response.status_code == 200
    data = response.get_json()
    assert len(data["data"]) == 1


def test_siswa_detail_export_returns_xlsx(logged_in_client, app):
    student_id = _make_student_with_requests(
        app, [("2025-01-10", "accepted")]
    )
    response = logged_in_client.get(
        "/admin/siswa/{}/export".format(student_id)
    )
    assert response.status_code == 200
    assert response.content_type == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert b"PK" in response.data


def test_siswa_detail_export_date_filter(logged_in_client, app):
    student_id = _make_student_with_requests(
        app,
        [("2025-01-10", "accepted"), ("2025-02-10", "pending")],
    )
    response = logged_in_client.get(
        "/admin/siswa/{}/export".format(student_id),
        query_string={"start_date": "2025-02-01", "end_date": "2025-02-28"},
    )
    assert response.status_code == 200
    import openpyxl
    from io import BytesIO

    wb = openpyxl.load_workbook(BytesIO(response.data))
    ws = wb.active
    # header at row 6, data at row 7 -> only one filtered request
    assert ws.cell(row=7, column=2).value == "10/02/2025"
    assert ws.cell(row=8, column=2).value is None


def test_siswa_detail_export_404(logged_in_client):
    response = logged_in_client.get("/admin/siswa/9999/export")
    assert response.status_code == 404


def test_siswa_detail_regular_admin_category_filter(
    regular_admin_client, app, regular_admin
):
    # siswa has requests in a category the regular admin does NOT manage
    student_id = _make_student_with_requests(
        app, [("2025-01-10", "accepted")]
    )
    response = regular_admin_client.get(
        "/admin/siswa/{}/data".format(student_id)
    )
    assert response.status_code == 200
    data = response.get_json()
    # regular admin manages no categories -> sees no requests
    assert len(data["data"]) == 0
