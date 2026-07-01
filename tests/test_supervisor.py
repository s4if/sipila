from datetime import date, datetime, timezone
from itertools import count

from app import db
from app.models import (
    BorrowingRequest,
    Category,
    ClassGroup,
    Student,
    Teacher,
)
from werkzeug.security import generate_password_hash

_setup_seq = count()


def _setup_borrowing_request(app, status="accepted", confirmation=None, tanggal=None):
    with app.app_context():
        n = next(_setup_seq)
        teacher = Teacher(
            username="guru_review_{}".format(n),
            password=generate_password_hash("pass"),
            name="Guru Review",
        )
        db.session.add(teacher)
        db.session.flush()

        cg = ClassGroup(name="1", grade_level="X", major="TJKT")
        db.session.add(cg)
        db.session.flush()

        student = Student(
            student_id="S{:03d}".format(n + 1),
            name="Siswa Satu",
            password=generate_password_hash("pass"),
            class_group_id=cg.id,
        )
        db.session.add(student)
        db.session.flush()

        cat = Category(name="Laptop {}".format(n))
        db.session.add(cat)
        db.session.flush()

        req = BorrowingRequest(
            student_id=student.id,
            category_id=cat.id,
            date=tanggal if tanggal else date.today(),
            status=status,
            reviewed_by=teacher.id,
            reviewed_at=datetime.now(timezone.utc),
            confirmation=confirmation,
        )
        if confirmation:
            req.confirmed_by = teacher.id
            req.confirmed_at = datetime.now(timezone.utc)
        db.session.add(req)
        db.session.commit()

        return req.id


def test_monitor_page_returns_200(logged_in_client):
    response = logged_in_client.get("/supervisor/monitor")
    assert response.status_code == 200
    assert b"Monitor Peminjaman" in response.data


def test_monitor_page_redirects_anonymous(client):
    response = client.get("/supervisor/monitor")
    assert response.status_code == 302
    assert "login" in response.location


def test_monitor_data_returns_json(logged_in_client, app):
    _setup_borrowing_request(app)
    response = logged_in_client.get("/supervisor/monitor/data")
    assert response.status_code == 200
    data = response.get_json()
    assert "data" in data
    assert len(data["data"]) == 1
    assert data["data"][0]["confirmation"] is None


def test_monitor_data_includes_confirmation_used(logged_in_client, app):
    _setup_borrowing_request(app, confirmation="used")
    response = logged_in_client.get("/supervisor/monitor/data")
    data = response.get_json()
    assert data["data"][0]["confirmation"] == "used"


def test_monitor_data_includes_confirmation_not_used(logged_in_client, app):
    _setup_borrowing_request(app, confirmation="not_used")
    response = logged_in_client.get("/supervisor/monitor/data")
    data = response.get_json()
    assert data["data"][0]["confirmation"] == "not_used"


def test_monitor_konfirmasi_used(logged_in_client, app):
    req_id = _setup_borrowing_request(app)
    response = logged_in_client.post(
        "/supervisor/monitor/konfirmasi/{}".format(req_id),
        data={"confirmation": "used"},
    )
    assert response.status_code == 200
    data = response.get_json()
    assert data["success"] is True
    assert "digunakan" in data["message"]

    with app.app_context():
        req = db.session.get(BorrowingRequest, req_id)
        assert req.confirmation == "used"
        assert req.confirmed_by is not None
        assert req.confirmed_at is not None


def test_monitor_konfirmasi_not_used(logged_in_client, app):
    req_id = _setup_borrowing_request(app)
    response = logged_in_client.post(
        "/supervisor/monitor/konfirmasi/{}".format(req_id),
        data={"confirmation": "not_used"},
    )
    assert response.status_code == 200
    data = response.get_json()
    assert data["success"] is True
    assert "tidak digunakan" in data["message"]

    with app.app_context():
        req = db.session.get(BorrowingRequest, req_id)
        assert req.confirmation == "not_used"


def test_monitor_konfirmasi_invalid_value(logged_in_client, app):
    req_id = _setup_borrowing_request(app)
    response = logged_in_client.post(
        "/supervisor/monitor/konfirmasi/{}".format(req_id),
        data={"confirmation": "invalid"},
    )
    assert response.status_code == 400
    data = response.get_json()
    assert data["success"] is False


def test_monitor_konfirmasi_rejected_request_fails(logged_in_client, app):
    req_id = _setup_borrowing_request(app, status="rejected")
    response = logged_in_client.post(
        "/supervisor/monitor/konfirmasi/{}".format(req_id),
        data={"confirmation": "used"},
    )
    assert response.status_code == 400
    data = response.get_json()
    assert data["success"] is False


def test_monitor_konfirmasi_pending_request_fails(logged_in_client, app):
    req_id = _setup_borrowing_request(app, status="pending")
    response = logged_in_client.post(
        "/supervisor/monitor/konfirmasi/{}".format(req_id),
        data={"confirmation": "used"},
    )
    assert response.status_code == 400
    data = response.get_json()
    assert data["success"] is False


def test_monitor_konfirmasi_404(logged_in_client):
    response = logged_in_client.post(
        "/supervisor/monitor/konfirmasi/9999",
        data={"confirmation": "used"},
    )
    assert response.status_code == 404


def test_monitor_konfirmasi_requires_login(client):
    response = client.post(
        "/supervisor/monitor/konfirmasi/1",
        data={"confirmation": "used"},
    )
    assert response.status_code == 302
    assert "login" in response.location


def test_monitor_batalkan_konfirmasi(logged_in_client, app):
    req_id = _setup_borrowing_request(app, confirmation="used")
    response = logged_in_client.post(
        "/supervisor/monitor/batalkan_konfirmasi/{}".format(req_id),
    )
    assert response.status_code == 200
    data = response.get_json()
    assert data["success"] is True
    assert "dibatalkan" in data["message"]

    with app.app_context():
        req = db.session.get(BorrowingRequest, req_id)
        assert req.confirmation is None
        assert req.confirmed_by is None
        assert req.confirmed_at is None


def test_monitor_batalkan_unconfirmed_fails(logged_in_client, app):
    req_id = _setup_borrowing_request(app)
    response = logged_in_client.post(
        "/supervisor/monitor/batalkan_konfirmasi/{}".format(req_id),
    )
    assert response.status_code == 400
    data = response.get_json()
    assert data["success"] is False


def test_monitor_batalkan_rejected_request_fails(logged_in_client, app):
    req_id = _setup_borrowing_request(app, status="rejected", confirmation="used")
    response = logged_in_client.post(
        "/supervisor/monitor/batalkan_konfirmasi/{}".format(req_id),
    )
    assert response.status_code == 400


def test_monitor_batalkan_404(logged_in_client):
    response = logged_in_client.post(
        "/supervisor/monitor/batalkan_konfirmasi/9999",
    )
    assert response.status_code == 404


def test_monitor_batalkan_requires_login(client):
    response = client.post("/supervisor/monitor/batalkan_konfirmasi/1")
    assert response.status_code == 302
    assert "login" in response.location


def test_monitor_konfirmasi_reversible(logged_in_client, app):
    req_id = _setup_borrowing_request(app)

    logged_in_client.post(
        "/supervisor/monitor/konfirmasi/{}".format(req_id),
        data={"confirmation": "used"},
    )
    with app.app_context():
        req = db.session.get(BorrowingRequest, req_id)
        assert req.confirmation == "used"

    logged_in_client.post(
        "/supervisor/monitor/batalkan_konfirmasi/{}".format(req_id),
    )
    with app.app_context():
        req = db.session.get(BorrowingRequest, req_id)
        assert req.confirmation is None

    logged_in_client.post(
        "/supervisor/monitor/konfirmasi/{}".format(req_id),
        data={"confirmation": "not_used"},
    )
    with app.app_context():
        req = db.session.get(BorrowingRequest, req_id)
        assert req.confirmation == "not_used"


def test_monitor_export_requires_login(client):
    response = client.get("/supervisor/monitor/export")
    assert response.status_code == 302
    assert "login" in response.location


def test_monitor_export_returns_xlsx(logged_in_client, app):
    _setup_borrowing_request(app)
    response = logged_in_client.get("/supervisor/monitor/export")
    assert response.status_code == 200
    assert (
        response.mimetype
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in response.headers["Content-Disposition"]
    assert "laporan_peminjaman" in response.headers["Content-Disposition"]


def test_monitor_export_contains_row(logged_in_client, app):
    import io
    import openpyxl

    _setup_borrowing_request(app)
    response = logged_in_client.get("/supervisor/monitor/export")
    wb = openpyxl.load_workbook(io.BytesIO(response.data))
    ws = wb.active
    # header pada baris 5, data mulai baris 6
    assert ws.cell(row=5, column=1).value == "No"
    assert ws.cell(row=6, column=4).value == "Siswa Satu"
    assert ws.cell(row=6, column=9).value == "Belum Dikonfirmasi"


def test_monitor_export_only_used_filter(logged_in_client, app):
    import io
    import openpyxl

    _setup_borrowing_request(app, confirmation=None)
    _setup_borrowing_request(app, confirmation="used")
    _setup_borrowing_request(app, confirmation="not_used")

    # semua permintaan
    resp_all = logged_in_client.get("/supervisor/monitor/export")
    wb_all = openpyxl.load_workbook(io.BytesIO(resp_all.data))
    # hitung baris data (header di baris 5)
    jumlah_semua = wb_all.active.max_row - 5
    assert jumlah_semua == 3

    # hanya yang digunakan
    resp_used = logged_in_client.get("/supervisor/monitor/export?hanya_digunakan=1")
    wb_used = openpyxl.load_workbook(io.BytesIO(resp_used.data))
    ws_used = wb_used.active
    jumlah_digunakan = ws_used.max_row - 5
    assert jumlah_digunakan == 1
    assert ws_used.cell(row=6, column=9).value == "Digunakan"
    assert ws_used["A3"].value == "Filter: Hanya yang digunakan"


def test_monitor_export_date_range(logged_in_client, app):
    import io
    import openpyxl

    _setup_borrowing_request(app, tanggal=date(2025, 1, 15))
    _setup_borrowing_request(app, tanggal=date(2025, 6, 10))

    # rentang hanya mencakup data pertama
    resp = logged_in_client.get(
        "/supervisor/monitor/export?tanggal_mulai=2025-01-01&tanggal_akhir=2025-02-01"
    )
    wb = openpyxl.load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    assert ws.max_row - 5 == 1
    assert ws.cell(row=6, column=2).value == "15/01/2025"
